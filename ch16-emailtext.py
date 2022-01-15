"""
### Sending Email and Text Messages
"""
"""
#
"""
"""
#
"""
"""
### Project: Sending Member Dues Reminder Emails

#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.
import openpyxl, smtplib, sys
# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_address@gmail.com', sys.argv[1])

# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not 
paid dues for %s. Please make this payment as soon as possible. Thank you!'" % 
(latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, 
        sendmailStatus))
smtpObj.quit()

"""
"""
#
"""
"""
#
"""
"""
### Sending Text with Twilio

"""
"""
### Project: "Just text me" Module
#! python3
# textMyself.py - Defines the textmyself() function that texts a message 
# passed to it as a string.

# Preset values:
accountSID = 'ACxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
authToken  = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
myNumber = '+15559998888'
twilioNumber = '+15552225678'

from twilio.rest import TwilioRestClient

def textmyself(message):
    twilioCli = TwilioRestClient(accountSID, authToken)
    twilioCli.messages.create(body=message, from_=twilioNumber, to=myNumber)


import textmyself
textmyself.textmyself('The boring task is finished.')
"""
###################
"""
### Practice Project: Random Chore Assignment Emailer
"""
"""
### Practice Project: Umbrella Reminder
"""
"""
### Practice Project: Auto Unsubscriber
"""
"""
### Practice Project: Controlling your computer through email
"""
"""
#
"""